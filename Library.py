import pandas as pd
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

class Publication(ABC):
    def __init__(self, title, author, year):
        self.title = title
        self.author = author
        self.year = year

    @abstractmethod
    def display(self):
        pass

class Book(Publication):
    def __init__(self, title, author, year, isbn, category, quantity):
        super().__init__(title, author, year)
        self.isbn = isbn
        self.category = category
        self.quantity = quantity
        self.available = quantity

    def display(self):
        print(f"Book: {self.title}, Author: {self.author}, Year: {self.year}, "
              f"ISBN: {self.isbn}, Category: {self.category}, "
              f"Available: {self.available}/{self.quantity}")

class Member:
    def __init__(self, name, member_id, contact):
        self.name = name
        self.member_id = member_id
        self.contact = contact

    def display(self):
        print(f"Member ID: {self.member_id}, Name: {self.name}, Contact: {self.contact}")

class Loan:
    def __init__(self, member, publication):
        self.member = member
        self.publication = publication
        self.loan_date = datetime.now()
        self.due_date = self.loan_date + timedelta(days=14)
        self.return_date = None

    def display(self):
        status = f"Due Date: {self.due_date.date()}" if not self.return_date else f"Returned on: {self.return_date.date()}"
        print(f"Loan - Member: {self.member.name}, Book: {self.publication.title}, Loan Date: {self.loan_date.date()}, {status}")

class Library:
    def __init__(self):
        self.members = []
        self.publications = []
        self.active_loans = []
        self.loan_history = []
        
        # Create data directory if it doesn't exist
        if not os.path.exists('data'):
            os.makedirs('data')
            
        # Load initial data
        self.load_books_from_file("books.xlsx")
        self.load_members_from_file("members.xlsx")
        self.load_loans_from_file("data/loans.xlsx")

    def add_member(self, member):
        if any(m.member_id == member.member_id for m in self.members):
            print(f"Member ID '{member.member_id}' already exists.")
            return False
        
        self.members.append(member)
        print(f"Member {member.name} added successfully.")
        self._save_members_to_file()
        return True

    def loan_publication(self, member_id, title):
        member = next((m for m in self.members if m.member_id == member_id), None)
        pub = next((p for p in self.publications if p.title.lower() == title.lower()), None)
        
        if not member:
            print(f"Member with ID '{member_id}' not found.")
            return False
        if not pub:
            print(f"Book '{title}' not found.")
            return False
        if pub.available <= 0:
            print(f"'{title}' is not available for loan.")
            return False
            
        loan = Loan(member, pub)
        self.active_loans.append(loan)
        self.loan_history.append(loan)
        pub.available -= 1
        
        # Save to files
        self._save_loans_to_file()
        self._save_books_to_file()
        
        print(f"{member.name} borrowed '{pub.title}'. Available copies: {pub.available}/{pub.quantity}")
        return True

    def return_publication(self, member_id, title):
        member = next((m for m in self.members if m.member_id == member_id), None)
        if not member:
            print(f"Member with ID '{member_id}' not found.")
            return False
            
        for loan in self.active_loans:
            if (loan.publication.title.lower() == title.lower() and 
                loan.member.member_id == member_id and 
                not loan.return_date):
                
                loan.return_date = datetime.now()
                loan.publication.available += 1
                self.active_loans.remove(loan)
                
                # Save to files
                self._save_loans_to_file()
                self._save_books_to_file()
                
                print(f"Return successful: {member.name} returned '{title}'. Available copies: {loan.publication.available}/{loan.publication.quantity}")
                return True
                
        print("Loan not found or book already returned.")
        return False

    def search_library(self, query):
        member = next((m for m in self.members if m.member_id == query), None)
        publication = next((p for p in self.publications if p.title.lower() == query.lower()), None)
        if member:
            member.display()
            print("\nCurrent Loans:")
            has_loans = False
            for loan in self.active_loans:
                if loan.member.member_id == member.member_id:
                    loan.display()
                    has_loans = True
            if not has_loans:
                print("No active loans for this member.")
        elif publication:
            publication.display()
            print("\nLoan Status:")
            active_loans = [loan for loan in self.active_loans 
                           if loan.publication.title.lower() == publication.title.lower()]
            if active_loans:
                print(f"Borrowed by:")
                for loan in active_loans:
                    print(f"- {loan.member.name} (ID: {loan.member.member_id}), Due: {loan.due_date.date()}")
            else:
                print("Available for borrowing")
        else:
            print("No matching book found.")

    def popular_books(self):
        print("\nPopular Books:")
        if not self.loan_history:
            print("No books have been borrowed yet.")
            return
        book_count = Counter([loan.publication.title for loan in self.loan_history])
        for title, count in book_count.most_common():
            print(f"- {title}: Borrowed {count} times")

    def show_members(self):
        print("\nMembers:")
        for member in self.members:
            member.display()

    def show_books(self):
        print("\nBooks:")
        for publication in self.publications:
            publication.display()

    def load_books_from_file(self, filename):
        try:
            if os.path.exists(filename):
                df = pd.read_excel(filename, engine='openpyxl')
                for _, row in df.iterrows():
                    try:
                        book = Book(
                            title=row['Title'],
                            author=row['Author'],
                            year=int(row['Year']),
                            isbn=row['ISBN'],
                            category=row['Category'],
                            quantity=int(row['Quantity'])
                        )
                        # Calculate available copies
                        active_loans = sum(1 for loan in self.active_loans 
                                          if loan.publication.title.lower() == row['Title'].lower())
                        book.available = book.quantity - active_loans
                        self.publications.append(book)
                    except ValueError as e:
                        print(f"Error processing book {row['Title']}: {e}")
        except Exception as e:
            print(f"An error occurred while reading the Excel file: {e}")

    def load_members_from_file(self, filename):
        try:
            if os.path.exists(filename):
                df = pd.read_excel(filename, engine='openpyxl')
                for _, row in df.iterrows():
                    member = Member(
                        name=row['Name'],
                        member_id=str(row['Member_ID']),
                        contact=row['Contact']
                    )
                    self.members.append(member)
        except Exception as e:
            print(f"An error occurred while reading the Excel file: {e}")

    def load_loans_from_file(self, filename):
        try:
            if os.path.exists(filename):
                df = pd.read_excel(filename, engine='openpyxl')
                for _, row in df.iterrows():
                    member = next((m for m in self.members if m.member_id == str(row['member_id'])), None)
                    publication = next((p for p in self.publications if p.title.lower() == row['book_title'].lower()), None)
                    
                    if member and publication:
                        loan = Loan(member, publication)
                        loan.loan_date = pd.to_datetime(row['loan_date'])
                        loan.due_date = pd.to_datetime(row['due_date'])
                        
                        if pd.notna(row['return_date']):
                            loan.return_date = pd.to_datetime(row['return_date'])
                        else:
                            self.active_loans.append(loan)
                            
                        self.loan_history.append(loan)
        except Exception as e:
            print(f"An error occurred while reading loans file: {e}")

    def _save_members_to_file(self):
        filename = 'members.xlsx'
        try:
            if not self.members:
                print("No members to save.")
                return
            data = []
            for member in self.members:
                data.append({
                    'Name': member.name,
                    'Member_ID': member.member_id,
                    'Contact': member.contact
                })
                
            df = pd.DataFrame(data)
            df.to_excel(filename, index=False, engine='openpyxl')
            print("Members saved successfully.")

            wb = load_workbook(filename)
            ws = wb.active
            column_widths = {"A": 20, "B": 18, "C": 25}  # กำหนดความกว้างคอลัมน์ตามต้องการ
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            wb.save(filename)
            print("Members saved successfully with adjusted column widths.")

        except Exception as e:
            print(f"Error saving members to file: {e}")

    def _save_books_to_file(self):
        try:
            data = []
            for book in self.publications:
                data.append({
                    'Title': book.title,
                    'Author': book.author,
                    'Year': book.year,
                    'ISBN': book.isbn,
                    'Category': book.category,
                    'Quantity': book.quantity,
                    'Available': book.available
                })
                
            df = pd.DataFrame(data)
            filenames = 'books.xlsx'
            df.to_excel(filenames, index=False)

            wb = load_workbook(filenames)
            ws = wb.active

            column_widths = {"A": 56, "B": 25, "C": 15, "D": 25, "E": 22, "F": 10, "G": 10}  # ปรับขนาดคอลัมน์
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            wb.save(filenames)
        except Exception as e:
            print(f"Error saving books to file: {e}")


    def _save_loans_to_file(self):
        try:
            data = []
            for loan in self.loan_history:
                data.append({
                    'member_id': loan.member.member_id,
                    'member_name': loan.member.name,
                    'book_title': loan.publication.title,
                    'book_author': loan.publication.author,
                    'loan_date': loan.loan_date,
                    'due_date': loan.due_date,
                    'return_date': loan.return_date if loan.return_date else None
                })
                
            df = pd.DataFrame(data)
            df.to_excel('data/loans.xlsx', index=False)
        except Exception as e:
            print(f"Error saving loans to file: {e}")

if __name__ == "__main__":
    lib = Library()

    while True:
        print("\nLibrary System")
        print("1. Add Member\n2. Loan Book\n3. Return Book")
        print("4. Search Library\n5. Popular Books\n6. Show Members")
        print("7. Show Books\n8. Exit")
        choice = input("Enter choice: ")

        if choice == "1":
            lib.add_member(Member(
                input("Enter name: "), 
                input("Enter ID: "), 
                input("Enter contact: ")
            ))
        elif choice == "2":
            lib.loan_publication(
                input("Enter Member ID: "), 
                input("Enter Book Title: ")
            )
        elif choice == "3":
            lib.return_publication(
                input("Enter Member ID: "), 
                input("Enter Book Title: ")
            )
        elif choice == "4":
            lib.search_library(input("Enter Book Title: "))
        elif choice == "5":
            lib.popular_books()
        elif choice == "6":
            lib.show_members()
        elif choice == "7":
            lib.show_books()
        elif choice == "8":
            # Save all data before exiting
            lib._save_members_to_file()
            lib._save_books_to_file()
            lib._save_loans_to_file()
            break
        else:
            print("Invalid choice.")
